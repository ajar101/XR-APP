"""
excel_builder.py — Engine pembuatan Excel multi-sheet.

Modul ini sepenuhnya bank-agnostik. Ia hanya membaca dua struktur data
standar yang dihasilkan oleh extractor manapun:

  saldo_per_bulan    → dict hasil BaseExtractor.extract_saldo()
  transaksi_per_bulan → dict hasil BaseExtractor.extract_transaksi()

Sheet yang dihasilkan:
  1. Saldo Harian
  2. Detail Transaksi
  3. Rekap Kredit
  4. Rekap Debit
  5. Cashflow Harian
  6. Kategori Debit
  7. Kategori Kredit
  8. Summary (+ HHI Score)
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from engine.categorizer import (
    kategorisasi_debit,
    kategorisasi_kredit,
    KATEGORI_DEBIT_KEYWORDS,
    KATEGORI_KREDIT_KEYWORDS,
)
from engine.anomaly_detector import detect_anomalies

BULAN_ORDER = [
    'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
    'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
]


# ============================================================
# HELPERS STYLING
# ============================================================

def style_header(cell, bg_color='1F4E79', font_color='FFFFFF'):
    cell.font      = Font(bold=True, color=font_color, size=11)
    cell.fill      = PatternFill(fill_type='solid', fgColor=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF'),
    )


def style_data(cell, align='left', bold=False, bg_color=None):
    cell.font      = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border    = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF'),
    )
    if bg_color:
        cell.fill = PatternFill(fill_type='solid', fgColor=bg_color)


def style_total_row(cell, align='right'):
    cell.font      = Font(bold=True, size=10)
    cell.fill      = PatternFill(fill_type='solid', fgColor='D9E1F2')
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border    = Border(
        left=Side(style='medium', color='1F4E79'),
        right=Side(style='medium', color='1F4E79'),
        top=Side(style='medium', color='1F4E79'),
        bottom=Side(style='medium', color='1F4E79'),
    )


def hitung_rata_rata_pengendapan(df) -> int:
    total_hari  = len(df)
    total_saldo = df['Saldo Akhir Harian'].sum()
    return int(total_saldo / total_hari) if total_hari else 0


# ============================================================
# ENTRY POINT
# ============================================================

def create_excel(saldo_per_bulan: dict, transaksi_per_bulan: dict,
                 output_path: str, bank_name: str = 'BANK', pdf_path: str = None) -> None:
    """
    Buat file Excel multi-sheet dari data saldo & transaksi.

    Args:
        saldo_per_bulan:     Output dari BaseExtractor.extract_saldo()
        transaksi_per_bulan: Output dari BaseExtractor.extract_transaksi()
        output_path:         Path file .xlsx yang akan disimpan
        bank_name:           Nama bank untuk label (misal 'BCA', 'MANDIRI')
        pdf_path:            Path PDF sumber — dipakai Sheet 9 (Indikasi Kejanggalan)
                              untuk pemeriksaan berbasis halaman mentah (running
                              balance, nomor halaman, metadata, dst). Opsional —
                              kalau tidak diberikan, sheet tetap dibuat tapi hanya
                              memuat pemeriksaan yang tidak butuh PDF mentah.
    """
    wb = Workbook()

    bulan_list = sorted(
        [b for b in saldo_per_bulan if not b.startswith('_')],
        key=lambda b: BULAN_ORDER.index(b) if b in BULAN_ORDER else 99
    )

    _build_sheet1_saldo(wb, saldo_per_bulan, bulan_list)
    _build_sheet2_transaksi(wb, transaksi_per_bulan, bulan_list)
    _build_sheet3_rekap_kredit(wb, transaksi_per_bulan, bulan_list)
    _build_sheet4_rekap_debit(wb, transaksi_per_bulan, bulan_list)
    _build_sheet5_cashflow(wb, saldo_per_bulan, transaksi_per_bulan, bulan_list)
    _build_sheet6_kategori_debit(wb, transaksi_per_bulan, bulan_list, saldo_per_bulan)
    _build_sheet7_kategori_kredit(wb, transaksi_per_bulan, bulan_list, saldo_per_bulan)
    _build_sheet8_summary(wb, saldo_per_bulan, transaksi_per_bulan, bulan_list, bank_name)
    _build_sheet9_indikasi(wb, saldo_per_bulan, transaksi_per_bulan, pdf_path)

    wb.save(output_path)


# ============================================================
# SHEET 1 — SALDO HARIAN
# ============================================================

def _build_sheet1_saldo(wb, saldo_per_bulan, bulan_list):
    ws = wb.active
    ws.title = 'Saldo Harian'
    ws.sheet_view.showGridLines = False

    col_offset = {0: 0, 1: 5, 2: 10}

    for idx, bulan in enumerate(bulan_list[:3]):
        info   = saldo_per_bulan[bulan]
        df     = info['df']
        offset = col_offset[idx]

        for h_idx, header in enumerate(['Bulan', 'Tanggal', 'Saldo Akhir Harian']):
            c = ws.cell(row=1, column=offset + h_idx + 1, value=header)
            style_header(c)

        for row_idx, row_data in df.iterrows():
            r  = row_idx + 2
            bg = 'F2F7FF' if row_idx % 2 == 0 else 'FFFFFF'

            c = ws.cell(row=r, column=offset + 1, value=row_data['Bulan'])
            style_data(c, align='center', bg_color=bg)

            c = ws.cell(row=r, column=offset + 2, value=row_data['Tanggal'])
            style_data(c, align='center', bg_color=bg)

            c = ws.cell(row=r, column=offset + 3, value=row_data['Saldo Akhir Harian'])
            c.number_format = '#,##0'
            style_data(c, align='right', bg_color=bg)

        rata_rata = hitung_rata_rata_pengendapan(df)
        c = ws.cell(row=34, column=offset + 1, value='Rata-rata Pengendapan')
        style_total_row(c, align='left')
        ws.cell(row=34, column=offset + 2).border = Border(
            top=Side(style='medium', color='1F4E79'),
            bottom=Side(style='medium', color='1F4E79'),
        )
        c = ws.cell(row=34, column=offset + 3, value=rata_rata)
        c.number_format = '#,##0'
        style_total_row(c, align='right')

        ws.column_dimensions[get_column_letter(offset + 1)].width = 22
        ws.column_dimensions[get_column_letter(offset + 2)].width = 10
        ws.column_dimensions[get_column_letter(offset + 3)].width = 22
        ws.row_dimensions[1].height = 30


# ============================================================
# SHEET 2 — DETAIL TRANSAKSI
# ============================================================

def _build_sheet2_transaksi(wb, transaksi_per_bulan, bulan_list):
    ws = wb.create_sheet(title='Detail Transaksi')
    ws.sheet_view.showGridLines = False

    col_offset = {0: 0, 1: 8, 2: 16}
    headers    = ['Bulan', 'Tanggal', 'Jenis Mutasi', 'Mutasi',
                  'Nama Pengirim/Penerima', 'Keterangan Transaksi']

    for idx, bulan in enumerate(bulan_list[:3]):
        if bulan not in transaksi_per_bulan:
            continue

        df     = transaksi_per_bulan[bulan]
        offset = col_offset[idx]

        for h_idx, header in enumerate(headers):
            c = ws.cell(row=1, column=offset + h_idx + 1, value=header)
            style_header(c)

        for row_idx, row_data in df.iterrows():
            r  = row_idx + 2
            bg = 'F2F7FF' if row_idx % 2 == 0 else 'FFFFFF'

            c = ws.cell(row=r, column=offset + 1, value=row_data['Bulan'])
            style_data(c, align='center', bg_color=bg)

            c = ws.cell(row=r, column=offset + 2, value=row_data['Tanggal'])
            style_data(c, align='center', bg_color=bg)

            jenis = row_data['Jenis Mutasi']
            c = ws.cell(row=r, column=offset + 3, value=jenis)
            style_data(c, align='center',
                       bg_color='E2EFDA' if jenis == 'Kredit' else 'FCE4EC')

            c = ws.cell(row=r, column=offset + 4, value=row_data['Mutasi'])
            c.number_format = '#,##0'
            style_data(c, align='right', bg_color=bg)

            c = ws.cell(row=r, column=offset + 5, value=row_data['Nama Pengirim/Penerima'])
            style_data(c, align='left', bg_color=bg)

            c = ws.cell(row=r, column=offset + 6, value=row_data['Keterangan Transaksi'])
            style_data(c, align='left', bg_color=bg)

        ws.column_dimensions[get_column_letter(offset + 1)].width = 12
        ws.column_dimensions[get_column_letter(offset + 2)].width = 8
        ws.column_dimensions[get_column_letter(offset + 3)].width = 13
        ws.column_dimensions[get_column_letter(offset + 4)].width = 18
        ws.column_dimensions[get_column_letter(offset + 5)].width = 26
        ws.column_dimensions[get_column_letter(offset + 6)].width = 52
        ws.row_dimensions[1].height = 30


# ============================================================
# HELPER — REKAP PIVOT (dipakai Sheet 3 & 4)
# ============================================================

def _build_rekap_sheet(ws, transaksi_per_bulan, bulan_list,
                       jenis_filter, label_total, show_concentration=False):
    ws.sheet_view.showGridLines = False

    bulan_ada = sorted(
        [b for b in bulan_list if b in transaksi_per_bulan],
        key=lambda b: BULAN_ORDER.index(b) if b in BULAN_ORDER else 99
    )

    frames = [transaksi_per_bulan[b][transaksi_per_bulan[b]['Jenis Mutasi'] == jenis_filter].copy()
              for b in bulan_ada]
    if not frames:
        return

    df_all = pd.concat(frames, ignore_index=True)

    pivot_nom = df_all.groupby(['Nama Pengirim/Penerima', 'Bulan'])['Mutasi'].sum().unstack(fill_value=0)
    pivot_nom = pivot_nom.reindex(columns=bulan_ada, fill_value=0)

    pivot_qty = df_all.groupby(['Nama Pengirim/Penerima', 'Bulan'])['Mutasi'].count().unstack(fill_value=0)
    pivot_qty = pivot_qty.reindex(columns=bulan_ada, fill_value=0)

    pivot_nom['Total'] = pivot_nom.sum(axis=1)
    pivot_qty['Total'] = pivot_qty.sum(axis=1)

    total_nom = {b: int(df_all[df_all['Bulan'] == b]['Mutasi'].sum()) for b in bulan_ada}
    total_qty = {b: int(df_all[df_all['Bulan'] == b]['Mutasi'].count()) for b in bulan_ada}
    total_nom['Total'] = int(df_all['Mutasi'].sum())
    total_qty['Total'] = int(df_all['Mutasi'].count())

    pivot_nom = pivot_nom.sort_values('Total', ascending=False)
    pivot_qty = pivot_qty.reindex(pivot_nom.index)

    pivot_nom = pivot_nom.replace(0, None)
    pivot_qty = pivot_qty.replace(0, None)

    grand_total = total_nom['Total']
    pct_list    = [(pivot_nom.loc[n, 'Total'] or 0) / grand_total * 100
                   for n in pivot_nom.index]
    pct_kum_list = []
    running = 0
    for p in pct_list:
        running += p
        pct_kum_list.append(running)

    # ---- Header ----
    ws.cell(row=1, column=1, value='Nama Pengirim/Penerima')
    style_header(ws.cell(row=1, column=1))
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30

    col = 2
    for b in bulan_ada:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        style_header(ws.cell(row=1, column=col, value=b))
        col += 2

    kolom_total_qty = col
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
    style_header(ws.cell(row=1, column=col, value='Total Qty'), bg_color='2E75B6')
    col += 1

    kolom_total_nom = col
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
    style_header(ws.cell(row=1, column=col, value=label_total), bg_color='2E75B6')
    col += 1

    kolom_pct = kolom_kum = None
    if show_concentration:
        kolom_pct = col
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        style_header(ws.cell(row=1, column=col, value='% Kontribusi'), bg_color='375623')
        col += 1
        kolom_kum = col
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        style_header(ws.cell(row=1, column=col, value='% Kumulatif'), bg_color='375623')

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    col = 2
    for b in bulan_ada:
        abbr = b[:3]
        style_header(ws.cell(row=2, column=col, value=f'{abbr} (Qty)'), bg_color='2E75B6')
        style_header(ws.cell(row=2, column=col + 1, value=f'{abbr} (Nominal)'), bg_color='2E75B6')
        col += 2

    # ---- Data rows ----
    for row_idx, (nama, nom_data) in enumerate(pivot_nom.iterrows()):
        r        = row_idx + 3
        bg       = 'F2F7FF' if row_idx % 2 == 0 else 'FFFFFF'
        qty_data = pivot_qty.loc[nama]

        c = ws.cell(row=r, column=1, value=nama)
        style_data(c, align='left', bg_color=bg)

        col = 2
        for b in bulan_ada:
            c = ws.cell(row=r, column=col, value=qty_data[b])
            style_data(c, align='center', bg_color=bg)
            c = ws.cell(row=r, column=col + 1, value=nom_data[b])
            c.number_format = '#,##0'
            style_data(c, align='right', bg_color=bg)
            col += 2

        c = ws.cell(row=r, column=kolom_total_qty, value=qty_data['Total'])
        style_data(c, align='center', bold=True, bg_color='EBF3FB')

        total_nom_val = nom_data['Total']
        c = ws.cell(row=r, column=kolom_total_nom, value=total_nom_val)
        c.number_format = '#,##0'
        style_data(c, align='right', bold=True, bg_color='EBF3FB')

        if show_concentration:
            pct     = (total_nom_val / grand_total * 100) if grand_total else 0
            pct_kum = pct_kum_list[row_idx]

            if pct_kum >= 80 and (row_idx == 0 or pct_kum_list[row_idx - 1] < 80):
                conc_bg = 'F4B942'
            elif pct_kum >= 50 and (row_idx == 0 or pct_kum_list[row_idx - 1] < 50):
                conc_bg = 'FFF176'
            else:
                conc_bg = bg

            c = ws.cell(row=r, column=kolom_pct, value=pct / 100)
            c.number_format = '0.0%'
            style_data(c, align='right', bg_color=conc_bg)

            c = ws.cell(row=r, column=kolom_kum, value=pct_kum / 100)
            c.number_format = '0.0%'
            style_data(c, align='right', bg_color=conc_bg)

    # ---- Baris total ----
    total_row = len(pivot_nom) + 3
    c = ws.cell(row=total_row, column=1, value=label_total)
    style_total_row(c, align='left')

    col = 2
    for b in bulan_ada:
        c = ws.cell(row=total_row, column=col, value=total_qty[b])
        style_total_row(c, align='center')
        c = ws.cell(row=total_row, column=col + 1, value=total_nom[b])
        c.number_format = '#,##0'
        style_total_row(c, align='right')
        col += 2

    c = ws.cell(row=total_row, column=kolom_total_qty, value=total_qty['Total'])
    style_total_row(c, align='center')
    c = ws.cell(row=total_row, column=kolom_total_nom, value=total_nom['Total'])
    c.number_format = '#,##0'
    style_total_row(c, align='right')

    if show_concentration:
        c = ws.cell(row=total_row, column=kolom_pct, value=1.0)
        c.number_format = '0.0%'
        style_total_row(c, align='right')
        style_total_row(ws.cell(row=total_row, column=kolom_kum))

    # ---- Lebar kolom ----
    ws.column_dimensions['A'].width = 28
    col = 2
    for _ in bulan_ada:
        ws.column_dimensions[get_column_letter(col)].width     = 10
        ws.column_dimensions[get_column_letter(col + 1)].width = 20
        col += 2
    ws.column_dimensions[get_column_letter(kolom_total_qty)].width = 12
    ws.column_dimensions[get_column_letter(kolom_total_nom)].width = 22
    if show_concentration:
        ws.column_dimensions[get_column_letter(kolom_pct)].width = 15
        ws.column_dimensions[get_column_letter(kolom_kum)].width = 15


def _build_sheet3_rekap_kredit(wb, transaksi_per_bulan, bulan_list):
    ws = wb.create_sheet(title='Rekap Kredit')
    _build_rekap_sheet(ws, transaksi_per_bulan, bulan_list,
                       'Kredit', 'Total Mutasi Kredit', show_concentration=True)


def _build_sheet4_rekap_debit(wb, transaksi_per_bulan, bulan_list):
    ws = wb.create_sheet(title='Rekap Debit')
    _build_rekap_sheet(ws, transaksi_per_bulan, bulan_list,
                       'Debit', 'Total Mutasi Debit')


# ============================================================
# SHEET 5 — CASHFLOW HARIAN
# ============================================================

def _build_sheet5_cashflow(wb, saldo_per_bulan, transaksi_per_bulan, bulan_list):
    ws = wb.create_sheet(title='Cashflow Harian')
    ws.sheet_view.showGridLines = False

    col_offset = {0: 0, 1: 7, 2: 14}
    headers    = ['Bulan', 'Tanggal', 'Total Kredit', 'Total Debit', 'Net Cashflow', 'Saldo Akhir']

    for idx, bulan in enumerate(bulan_list[:3]):
        if bulan not in transaksi_per_bulan:
            continue

        df       = transaksi_per_bulan[bulan]
        saldo_df = saldo_per_bulan[bulan]['df']
        offset   = col_offset[idx]

        for h_idx, header in enumerate(headers):
            c = ws.cell(row=1, column=offset + h_idx + 1, value=header)
            style_header(c)

        kredit_per_tgl = df[df['Jenis Mutasi'] == 'Kredit'].groupby('Tanggal')['Mutasi'].sum()
        debit_per_tgl  = df[df['Jenis Mutasi'] == 'Debit'].groupby('Tanggal')['Mutasi'].sum()
        tanggal_ada    = sorted(df['Tanggal'].unique())
        saldo_lookup   = dict(zip(saldo_df['Tanggal'], saldo_df['Saldo Akhir Harian']))

        total_kredit = total_debit = 0

        for row_idx, tgl in enumerate(tanggal_ada):
            r  = row_idx + 2
            bg = 'F2F7FF' if row_idx % 2 == 0 else 'FFFFFF'

            kredit = int(kredit_per_tgl.get(tgl, 0))
            debit  = int(debit_per_tgl.get(tgl, 0))
            net    = kredit - debit
            saldo  = saldo_lookup.get(tgl)
            total_kredit += kredit
            total_debit  += debit

            c = ws.cell(row=r, column=offset + 1, value=bulan)
            style_data(c, align='center', bg_color=bg)
            c = ws.cell(row=r, column=offset + 2, value=tgl)
            style_data(c, align='center', bg_color=bg)
            c = ws.cell(row=r, column=offset + 3, value=kredit if kredit else None)
            c.number_format = '#,##0'
            style_data(c, align='right', bg_color=bg)
            c = ws.cell(row=r, column=offset + 4, value=debit if debit else None)
            c.number_format = '#,##0'
            style_data(c, align='right', bg_color=bg)
            net_bg = 'E2EFDA' if net >= 0 else 'FCE4EC'
            c = ws.cell(row=r, column=offset + 5, value=net)
            c.number_format = '#,##0'
            style_data(c, align='right', bg_color=net_bg)
            c = ws.cell(row=r, column=offset + 6, value=saldo)
            c.number_format = '#,##0'
            style_data(c, align='right', bg_color=bg)

        total_row = len(tanggal_ada) + 2
        total_net = total_kredit - total_debit

        for col_offset_i, val, fmt in [
            (1, bulan,        None),
            (3, total_kredit, '#,##0'),
            (4, total_debit,  '#,##0'),
            (5, total_net,    '#,##0'),
            (6, None,         None),
        ]:
            c = ws.cell(row=total_row, column=offset + col_offset_i, value=val)
            if fmt:
                c.number_format = fmt
            style_total_row(c, align='right' if fmt else 'left')

        ws.column_dimensions[get_column_letter(offset + 1)].width = 14
        ws.column_dimensions[get_column_letter(offset + 2)].width = 10
        ws.column_dimensions[get_column_letter(offset + 3)].width = 18
        ws.column_dimensions[get_column_letter(offset + 4)].width = 18
        ws.column_dimensions[get_column_letter(offset + 5)].width = 18
        ws.column_dimensions[get_column_letter(offset + 6)].width = 18
        ws.row_dimensions[1].height = 30


# ============================================================
# HELPER — KATEGORI SHEET (dipakai Sheet 6 & 7)
# ============================================================

def _build_kategori_sheet(ws, df_all, kategori_keys, label_total,
                          jenis_bg, bulan_ada):
    """Bangun sheet kategori (detail + pivot) untuk debit atau kredit."""
    ws.sheet_view.showGridLines = False

    disclaimer = (
        "CATATAN PENTING KATEGORISASI OTOMATIS: Sistem melakukan kategorisasi transaksi berdasarkan "
        "keyword yang terdeteksi di kolom Keterangan Transaksi dan Nama Pengirim/Penerima. Transaksi "
        "yang tidak memiliki keyword spesifik akan masuk ke kategori default. Mohon lakukan review "
        "manual untuk memastikan akurasi kategorisasi sesuai konteks bisnis. Anda dapat mengubah "
        "kategori secara langsung di file Excel ini."
    )
    ws.merge_cells('A1:G1')
    c = ws.cell(row=1, column=1, value=disclaimer)
    c.font      = Font(size=10, italic=True, color='D97706')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.fill      = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
    c.border    = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'),  bottom=Side(style='thin'))
    ws.row_dimensions[1].height = 60

    headers = ['Bulan', 'Tanggal', 'Jenis Mutasi', 'Mutasi',
               'Nama Pengirim/Penerima', 'Keterangan Transaksi', 'Kategori']
    for col_idx, header in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx, value=header)
        style_header(c)
    ws.row_dimensions[2].height = 30

    for idx, row_data in df_all.iterrows():
        r  = idx + 3
        bg = 'F2F7FF' if idx % 2 == 0 else 'FFFFFF'

        c = ws.cell(row=r, column=1, value=row_data['Bulan'])
        style_data(c, align='center', bg_color=bg)
        c = ws.cell(row=r, column=2, value=row_data['Tanggal'])
        style_data(c, align='center', bg_color=bg)
        c = ws.cell(row=r, column=3, value=row_data['Jenis Mutasi'])
        style_data(c, align='center', bg_color=jenis_bg)
        c = ws.cell(row=r, column=4, value=row_data['Mutasi'])
        c.number_format = '#,##0'
        style_data(c, align='right', bg_color=bg)
        c = ws.cell(row=r, column=5, value=row_data['Nama Pengirim/Penerima'])
        style_data(c, align='left', bg_color=bg)
        c = ws.cell(row=r, column=6, value=row_data['Keterangan Transaksi'])
        style_data(c, align='left', bg_color=bg)
        c = ws.cell(row=r, column=7, value=row_data['Kategori'])
        style_data(c, align='left', bold=True, bg_color=bg)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 28

    # ---- PIVOT TABLE (di samping kanan, mulai kolom 9) ----
    psc = 9   # pivot start column
    psr = 3   # pivot start row

    pivot_data = {}
    for kategori in kategori_keys:
        pivot_data[kategori] = {}
        for bulan in bulan_ada:
            subset = df_all[(df_all['Kategori'] == kategori) & (df_all['Bulan'] == bulan)]
            qty = len(subset)
            nom = int(subset['Mutasi'].sum()) if qty > 0 else 0
            pivot_data[kategori][bulan] = {'qty': qty, 'nom': nom}
        pivot_data[kategori]['Total'] = {
            'qty': sum(v['qty'] for v in pivot_data[kategori].values()),
            'nom': sum(v['nom'] for v in pivot_data[kategori].values()),
        }

    # Sub-header
    c = ws.cell(row=psr - 1, column=psc, value='Summary per Kategori')
    style_header(c)
    ws.merge_cells(start_row=psr - 1, start_column=psc,
                   end_row=psr - 1, end_column=psc + len(bulan_ada) * 2 + 2)
    ws.row_dimensions[psr - 1].height = 25

    # Header pivot row 1
    c = ws.cell(row=psr, column=psc, value='Kategori')
    style_header(c)
    ws.merge_cells(start_row=psr, start_column=psc, end_row=psr + 1, end_column=psc)

    col = psc + 1
    for bulan in bulan_ada:
        ws.merge_cells(start_row=psr, start_column=col, end_row=psr, end_column=col + 1)
        style_header(ws.cell(row=psr, column=col, value=bulan))
        col += 2

    for label_col, label_val in [('Total Qty', None), ('Total Nominal', None), ('% dari Total', None)]:
        ws.merge_cells(start_row=psr, start_column=col, end_row=psr + 1, end_column=col)
        style_header(ws.cell(row=psr, column=col, value=label_col), bg_color='2E75B6')
        col += 1

    # Header pivot row 2
    col = psc + 1
    for bulan in bulan_ada:
        abbr = bulan[:3]
        style_header(ws.cell(row=psr + 1, column=col,     value=f'{abbr} (Qty)'),    bg_color='2E75B6')
        style_header(ws.cell(row=psr + 1, column=col + 1, value=f'{abbr} (Nominal)'), bg_color='2E75B6')
        col += 2

    ws.row_dimensions[psr].height     = 20
    ws.row_dimensions[psr + 1].height = 30

    grand_total   = df_all['Mutasi'].sum()
    pivot_row_start = psr + 2
    kolom_total_qty = psc + 1 + len(bulan_ada) * 2
    kolom_total_nom = kolom_total_qty + 1
    kolom_pct       = kolom_total_nom + 1

    for row_idx, kategori in enumerate(kategori_keys):
        r  = pivot_row_start + row_idx
        bg = 'F2F7FF' if row_idx % 2 == 0 else 'FFFFFF'

        c = ws.cell(row=r, column=psc, value=kategori)
        style_data(c, align='left', bg_color=bg)

        col = psc + 1
        for bulan in bulan_ada:
            qty = pivot_data[kategori][bulan]['qty']
            nom = pivot_data[kategori][bulan]['nom']
            c = ws.cell(row=r, column=col,     value=qty if qty > 0 else None)
            style_data(c, align='center', bg_color=bg)
            c = ws.cell(row=r, column=col + 1, value=nom if nom > 0 else None)
            c.number_format = '#,##0'
            style_data(c, align='right', bg_color=bg)
            col += 2

        tqv = pivot_data[kategori]['Total']['qty']
        tnv = pivot_data[kategori]['Total']['nom']
        pct = (tnv / grand_total * 100) if grand_total > 0 else 0

        c = ws.cell(row=r, column=kolom_total_qty, value=tqv if tqv > 0 else None)
        style_data(c, align='center', bold=True, bg_color='EBF3FB')
        c = ws.cell(row=r, column=kolom_total_nom, value=tnv if tnv > 0 else None)
        c.number_format = '#,##0'
        style_data(c, align='right', bold=True, bg_color='EBF3FB')
        c = ws.cell(row=r, column=kolom_pct, value=pct / 100 if pct > 0 else None)
        c.number_format = '0.0%'
        style_data(c, align='right', bold=True, bg_color='EBF3FB')

    # Baris total pivot
    total_row = pivot_row_start + len(kategori_keys)
    c = ws.cell(row=total_row, column=psc, value=label_total)
    style_total_row(c, align='left')

    col = psc + 1
    for bulan in bulan_ada:
        tqb = sum(pivot_data[k][bulan]['qty'] for k in pivot_data)
        tnb = sum(pivot_data[k][bulan]['nom'] for k in pivot_data)
        c = ws.cell(row=total_row, column=col, value=tqb)
        style_total_row(c, align='center')
        c = ws.cell(row=total_row, column=col + 1, value=tnb)
        c.number_format = '#,##0'
        style_total_row(c, align='right')
        col += 2

    c = ws.cell(row=total_row, column=kolom_total_qty, value=len(df_all))
    style_total_row(c, align='center')
    c = ws.cell(row=total_row, column=kolom_total_nom, value=int(grand_total))
    c.number_format = '#,##0'
    style_total_row(c, align='right')
    c = ws.cell(row=total_row, column=kolom_pct, value=1.0)
    c.number_format = '0.0%'
    style_total_row(c, align='right')

    for col_i in range(psc, psc + len(bulan_ada) * 2 + 4):
        ws.column_dimensions[get_column_letter(col_i)].width = 18


def _build_sheet6_kategori_debit(wb, transaksi_per_bulan, bulan_list, saldo_per_bulan):
    ws = wb.create_sheet(title='Kategori Debit')

    nama_perusahaan = saldo_per_bulan.get('_nama_pemilik', '')
    bulan_ada = sorted(
        [b for b in bulan_list if b in transaksi_per_bulan],
        key=lambda b: BULAN_ORDER.index(b) if b in BULAN_ORDER else 99
    )

    frames = []
    for bulan in bulan_ada:
        df_b = transaksi_per_bulan[bulan][
            transaksi_per_bulan[bulan]['Jenis Mutasi'] == 'Debit'
        ].copy()
        if not df_b.empty:
            df_b['Kategori'] = df_b.apply(
                lambda r: kategorisasi_debit(
                    r.get('Keterangan Transaksi', ''),
                    r.get('Nama Pengirim/Penerima', ''),
                    r.get('Mutasi', 0),
                    nama_perusahaan
                ), axis=1
            )
        else:
            df_b['Kategori'] = pd.Series(dtype='object')
        frames.append(df_b)

    if not frames:
        return

    df_all = pd.concat(frames, ignore_index=True)
    kategori_keys = list(KATEGORI_DEBIT_KEYWORDS.keys()) + ['Lain-lain Tanpa Keterangan']
    _build_kategori_sheet(ws, df_all, kategori_keys, 'Total Debit', 'FCE4EC', bulan_ada)


def _build_sheet7_kategori_kredit(wb, transaksi_per_bulan, bulan_list, saldo_per_bulan):
    ws = wb.create_sheet(title='Kategori Kredit')

    nama_perusahaan = saldo_per_bulan.get('_nama_pemilik', '')
    bulan_ada = sorted(
        [b for b in bulan_list if b in transaksi_per_bulan],
        key=lambda b: BULAN_ORDER.index(b) if b in BULAN_ORDER else 99
    )

    frames = []
    for bulan in bulan_ada:
        df_b = transaksi_per_bulan[bulan][
            transaksi_per_bulan[bulan]['Jenis Mutasi'] == 'Kredit'
        ].copy()
        if not df_b.empty:
            df_b['Kategori'] = df_b.apply(
                lambda r: kategorisasi_kredit(
                    r.get('Keterangan Transaksi', ''),
                    r.get('Nama Pengirim/Penerima', ''),
                    r.get('Mutasi', 0),
                    nama_perusahaan
                ), axis=1
            )
        else:
            df_b['Kategori'] = pd.Series(dtype='object')
        frames.append(df_b)

    if not frames:
        return

    df_all = pd.concat(frames, ignore_index=True)
    kategori_keys = list(KATEGORI_KREDIT_KEYWORDS.keys()) + ['Lain-lain']
    _build_kategori_sheet(ws, df_all, kategori_keys, 'Total Kredit', 'E2EFDA', bulan_ada)


# ============================================================
# SHEET 8 — SUMMARY + HHI SCORE
# ============================================================

def _build_sheet8_summary(wb, saldo_per_bulan, transaksi_per_bulan,
                           bulan_list, bank_name):
    ws = wb.create_sheet(title='Summary')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 2

    nama_pemilik = saldo_per_bulan.get('_nama_pemilik', '-')
    no_rekening  = saldo_per_bulan.get('_no_rekening', '-')

    bulan_summary = sorted(
        [b for b in bulan_list if b in transaksi_per_bulan],
        key=lambda b: BULAN_ORDER.index(b) if b in BULAN_ORDER else 99
    )
    n_bulan = len(bulan_summary)

    for i in range(n_bulan + 2):
        ws.column_dimensions[get_column_letter(3 + i)].width = 22

    # ---- Disclaimer ----
    disclaimer = (
        "PERHATIAN: XR-App membantu Anda menganalisis rekening koran secara otomatis, namun tidak "
        "menjamin keakuratan 100%. Kemungkinan terdapat kesalahan ekstraksi data, salah kategorisasi, "
        "atau ketidaksesuaian lainnya. Mohon selalu lakukan cross-check dengan dokumen asli dan review "
        "manual sebelum menggunakan hasil ini untuk keputusan penting. Pengguna bertanggung jawab penuh "
        "atas validasi dan penggunaan data yang dihasilkan."
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=3 + n_bulan + 1)
    c = ws.cell(row=1, column=1, value=disclaimer)
    c.font      = Font(size=10, italic=True, bold=True, color='DC2626')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.fill      = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    c.border    = Border(left=Side(style='medium'), right=Side(style='medium'),
                         top=Side(style='medium'),  bottom=Side(style='medium'))
    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 15

    # ---- Identitas ----
    def write_identitas(row, label, value):
        c = ws.cell(row=row, column=1, value=label)
        c.font      = Font(bold=True, size=11)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c = ws.cell(row=row, column=3, value=f': {value}')
        c.font      = Font(size=11)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 20

    write_identitas(4, 'Bank',                   bank_name)
    write_identitas(5, 'Nomor Rekening',          no_rekening)
    write_identitas(6, 'Nama Pemilik Rekening',   nama_pemilik)

    # ---- Tabel ringkasan keuangan ----
    tbl_start = 8
    ws.row_dimensions[tbl_start].height = 28
    style_header(ws.cell(row=tbl_start, column=1, value=''))

    col = 3
    for bulan in bulan_summary:
        style_header(ws.cell(row=tbl_start, column=col, value=bulan))
        col += 1
    style_header(ws.cell(row=tbl_start, column=col, value='Total'), bg_color='2E75B6')
    col += 1
    style_header(ws.cell(row=tbl_start, column=col, value='Rata-rata'), bg_color='2E75B6')

    data_bulan = {}
    for bulan in bulan_summary:
        df_t   = transaksi_per_bulan[bulan]
        df_s   = saldo_per_bulan[bulan]['df']
        kredit = int(df_t[df_t['Jenis Mutasi'] == 'Kredit']['Mutasi'].sum())
        debit  = int(df_t[df_t['Jenis Mutasi'] == 'Debit']['Mutasi'].sum())
        rata   = hitung_rata_rata_pengendapan(df_s)
        if not df_s.empty:
            saldo_akhir = int(df_s['Saldo Akhir Harian'].iloc[-1])
            saldo_min   = int(df_s['Saldo Akhir Harian'].min())
            saldo_max   = int(df_s['Saldo Akhir Harian'].max())
        else:
            saldo_akhir = saldo_min = saldo_max = 0

        data_bulan[bulan] = {
            'Saldo Awal'            : saldo_per_bulan.get(f'_saldo_awal_{bulan}'),
            'Total Kredit'          : kredit,
            'Total Debit'           : debit,
            'Net Cashflow'          : kredit - debit,
            'Rata-rata Pengendapan' : rata,
            'Saldo Akhir'           : saldo_akhir,
            'Saldo Minimum'         : saldo_min,
            'Saldo Maksimum'        : saldo_max,
        }

    rows_label = ['Saldo Awal', 'Total Kredit', 'Total Debit', 'Net Cashflow',
                  'Rata-rata Pengendapan', 'Saldo Akhir', 'Saldo Minimum', 'Saldo Maksimum']
    bold_labels = {'Total Kredit', 'Total Debit', 'Net Cashflow', 'Rata-rata Pengendapan'}

    for r_idx, label in enumerate(rows_label):
        r  = tbl_start + 1 + r_idx
        bg = 'F2F7FF' if r_idx % 2 == 0 else 'FFFFFF'
        ws.row_dimensions[r].height = 20

        c = ws.cell(row=r, column=1, value=label)
        style_data(c, align='left', bold=(label in bold_labels), bg_color=bg)

        values = []
        col = 3
        for bulan in bulan_summary:
            val = data_bulan[bulan].get(label)
            values.append(val)
            c = ws.cell(row=r, column=col, value=val)
            c.number_format = '#,##0'
            cell_bg = ('E2EFDA' if val and val >= 0 else 'FCE4EC') if label == 'Net Cashflow' else bg
            style_data(c, align='right', bg_color=cell_bg)
            col += 1

        valid     = [v for v in values if v is not None]
        total_val = sum(valid) if valid else None
        avg_val   = int(total_val / len(valid)) if valid else None

        c = ws.cell(row=r, column=col, value=total_val)
        c.number_format = '#,##0'
        cell_bg = ('E2EFDA' if total_val and total_val >= 0 else 'FCE4EC') if label == 'Net Cashflow' else 'EBF3FB'
        style_data(c, align='right', bold=True, bg_color=cell_bg)

        c = ws.cell(row=r, column=col + 1, value=avg_val)
        c.number_format = '#,##0'
        cell_bg = ('E2EFDA' if avg_val and avg_val >= 0 else 'FCE4EC') if label == 'Net Cashflow' else 'EBF3FB'
        style_data(c, align='right', bold=True, bg_color=cell_bg)

    # ---- Konsentrasi kredit ----
    konsen_start = tbl_start + len(rows_label) + 3
    ws.row_dimensions[konsen_start - 1].height = 8

    c = ws.cell(row=konsen_start, column=1, value='Konsentrasi Kredit')
    style_header(c)
    ws.merge_cells(start_row=konsen_start, start_column=1,
                   end_row=konsen_start, end_column=3 + n_bulan)
    ws.row_dimensions[konsen_start].height = 25

    frames_cr = [transaksi_per_bulan[b][transaksi_per_bulan[b]['Jenis Mutasi'] == 'Kredit'].copy()
                 for b in bulan_summary]
    df_cr    = pd.concat(frames_cr, ignore_index=True)
    rekap_cr = df_cr.groupby('Nama Pengirim/Penerima')['Mutasi'].sum().sort_values(ascending=False)
    total_cr = rekap_cr.sum()
    n_aktif  = len(rekap_cr)
    pct_kum  = (rekap_cr.cumsum() / total_cr * 100)
    n_50     = int((pct_kum < 50).sum()) + 1
    n_80     = int((pct_kum < 80).sum()) + 1

    konsen_data = [
        ('Total Pengirim Aktif',          n_aktif),
        ('Pengirim mencapai 50% pertama', f'{n_50} dari {n_aktif}'),
        ('Pengirim mencapai 80% pertama', f'{n_80} dari {n_aktif}'),
    ]
    for k_idx, (label, value) in enumerate(konsen_data):
        r  = konsen_start + 1 + k_idx
        bg = 'F2F7FF' if k_idx % 2 == 0 else 'FFFFFF'
        ws.row_dimensions[r].height = 20
        c = ws.cell(row=r, column=1, value=label)
        style_data(c, align='left', bg_color=bg)
        c = ws.cell(row=r, column=3, value=value)
        style_data(c, align='left', bold=True, bg_color=bg)

    # ---- HHI Score ----
    if not rekap_cr.empty:
        pct_each = (rekap_cr / total_cr * 100)
        hhi      = int(sum(pct_each ** 2))
        top_nama = rekap_cr.index[0]
        top_pct  = round(pct_each.iloc[0], 1)
    else:
        hhi      = 0
        top_nama = "-"
        top_pct  = 0

    if hhi < 1500:
        hhi_kategori, hhi_bg, hhi_font_col = 'Diversified',   'E2EFDA', '375623'
    elif hhi <= 2500:
        hhi_kategori, hhi_bg, hhi_font_col = 'Moderate',      'FFF2CC', '7D6608'
    else:
        hhi_kategori, hhi_bg, hhi_font_col = 'Concentrated',  'FCE4EC', 'C00000'

    hhi_start = konsen_start + len(konsen_data) + 2
    ws.row_dimensions[hhi_start - 1].height = 8

    c = ws.cell(row=hhi_start, column=1, value='HHI Score (Herfindahl-Hirschman Index)')
    style_header(c)
    ws.merge_cells(start_row=hhi_start, start_column=1,
                   end_row=hhi_start, end_column=3 + n_bulan)
    ws.row_dimensions[hhi_start].height = 25

    def _hhi_row(row, label, value, value_size=11):
        ws.row_dimensions[row].height = 22
        c = ws.cell(row=row, column=1, value=label)
        c.font      = Font(bold=True, size=11)
        c.fill      = PatternFill(fill_type='solid', fgColor=hhi_bg)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border    = Border(left=Side(style='thin', color='BFBFBF'),
                             right=Side(style='thin', color='BFBFBF'),
                             top=Side(style='thin', color='BFBFBF'),
                             bottom=Side(style='thin', color='BFBFBF'))
        c = ws.cell(row=row, column=3, value=value)
        c.font      = Font(bold=True, size=value_size, color=hhi_font_col)
        c.fill      = PatternFill(fill_type='solid', fgColor=hhi_bg)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border    = Border(left=Side(style='thin', color='BFBFBF'),
                             right=Side(style='thin', color='BFBFBF'),
                             top=Side(style='thin', color='BFBFBF'),
                             bottom=Side(style='thin', color='BFBFBF'))
        if isinstance(value, int):
            c.number_format = '#,##0'

    _hhi_row(hhi_start + 1, 'HHI Score', hhi,          value_size=14)
    _hhi_row(hhi_start + 2, 'Kategori',  hhi_kategori, value_size=12)

    # ---- Panduan Interpretasi ----
    panduan_start = hhi_start + 4
    ws.row_dimensions[panduan_start - 1].height = 8

    c = ws.cell(row=panduan_start, column=1, value='Panduan Interpretasi HHI')
    style_header(c)
    ws.merge_cells(start_row=panduan_start, start_column=1,
                   end_row=panduan_start, end_column=3 + n_bulan)
    ws.row_dimensions[panduan_start].height = 25

    teks_div = (
        f"HHI < 1.500 (Diversified): Arus kas kredit berasal dari banyak sumber yang "
        f"terdiversifikasi ({n_aktif} pengirim aktif). Risiko kegagalan bayar akibat kehilangan "
        f"satu pelanggan sangat rendah."
    )
    teks_mod = (
        f"1.500 \u2264 HHI \u2264 2.500 (Moderate): Terdapat ketergantungan pada {n_80} pelanggan "
        f"utama ({n_80} pengirim menyumbang 80% arus kas kredit dari {n_aktif} pengirim aktif). "
        f"Lakukan verifikasi terhadap profil dan stabilitas pelanggan utama."
    )
    teks_con = (
        f"HHI > 2.500 (Concentrated): RISIKO TINGGI. Arus kas kredit terkonsentrasi pada "
        f"sedikit pihak \u2014 {n_80} pengirim pertama menguasai 80% total kredit masuk, "
        f"dengan pengirim terbesar {top_nama} berkontribusi {top_pct}%."
    )

    panduan_items = [
        ('E2EFDA', '375623', teks_div),
        ('FFF2CC', '7D6608', teks_mod),
        ('FCE4EC', 'C00000', teks_con),
    ]

    for p_idx, (p_bg, p_font, teks) in enumerate(panduan_items):
        r_p = panduan_start + 1 + p_idx
        ws.row_dimensions[r_p].height = 60
        is_active = (
            (p_idx == 0 and hhi_kategori == 'Diversified') or
            (p_idx == 1 and hhi_kategori == 'Moderate') or
            (p_idx == 2 and hhi_kategori == 'Concentrated')
        )
        border_style = 'medium' if is_active else 'thin'
        border_color = '1F4E79' if is_active else 'BFBFBF'

        ws.merge_cells(start_row=r_p, start_column=1, end_row=r_p, end_column=2 + n_bulan)
        c = ws.cell(row=r_p, column=1, value=teks)
        c.font      = Font(bold=is_active, size=10,
                           color=p_font if is_active else '595959')
        c.fill      = PatternFill(fill_type='solid',
                                   fgColor=p_bg if is_active else 'F9F9F9')
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border    = Border(
            left=Side(style=border_style, color=border_color),
            right=Side(style=border_style, color=border_color),
            top=Side(style=border_style, color=border_color),
            bottom=Side(style=border_style, color=border_color),
        )

    # ---- Signature ----
    sig_start = panduan_start + len(panduan_items) + 3
    ws.row_dimensions[sig_start - 1].height = 12
    n_cols_sig = 2 + n_bulan

    for sig_row, sig_val, sig_font, sig_border in [
        (sig_start,     'XR-App \u2014 eXtract-Report v1.0',
         Font(bold=True, size=11, color='1F4E79'),
         Border(top=Side(style='medium', color='1F4E79'))),
        (sig_start + 1, 'Automated Bank Statement Analysis Tool',
         Font(size=10, color='595959', italic=True), None),
        (sig_start + 2, '\u00a9 2026 Ajar D Ashaq. All rights reserved.',
         Font(size=9, color='595959'),
         Border(bottom=Side(style='medium', color='1F4E79'))),
    ]:
        ws.merge_cells(start_row=sig_row, start_column=1,
                       end_row=sig_row, end_column=1 + n_cols_sig)
        c = ws.cell(row=sig_row, column=1, value=sig_val)
        c.font      = sig_font
        c.fill      = PatternFill(fill_type='solid', fgColor='EBF3FB')
        c.alignment = Alignment(horizontal='center', vertical='center')
        if sig_border:
            c.border = sig_border
        ws.row_dimensions[sig_row].height = 20 if sig_row == sig_start else 18


# ============================================================
# SHEET 9 — INDIKASI KEJANGGALAN
# ============================================================

TINGKAT_STYLE = {
    'Tinggi': ('FCE4EC', 'C00000'),
    'Sedang': ('FFF2CC', '7D6608'),
    'Rendah': ('E2EFDA', '375623'),
}


def _build_sheet9_indikasi(wb, saldo_per_bulan, transaksi_per_bulan, pdf_path):
    ws = wb.create_sheet(title='Indikasi Kejanggalan')
    ws.sheet_view.showGridLines = False

    widths = {'A': 4, 'B': 26, 'C': 10, 'D': 10, 'E': 8, 'F': 8,
              'G': 46, 'H': 55, 'I': 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    findings = detect_anomalies(pdf_path, saldo_per_bulan, transaksi_per_bulan) if pdf_path else []

    # ---- Disclaimer ----
    disclaimer = (
        "PERHATIAN: Sheet ini berisi indikasi otomatis yang PERLU DIVERIFIKASI MANUAL, "
        "bukan kesimpulan akhir bahwa rekening tidak wajar. Beberapa temuan (mis. transaksi bulat "
        "atau berulang) bisa jadi wajar tergantung profil nasabah. Gunakan sebagai titik awal review, "
        "bukan pengganti audit atau uji kelayakan (due diligence) manual."
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=9)
    c = ws.cell(row=1, column=1, value=disclaimer)
    c.font      = Font(size=10, italic=True, bold=True, color='DC2626')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.fill      = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    c.border    = Border(left=Side(style='medium'), right=Side(style='medium'),
                         top=Side(style='medium'),  bottom=Side(style='medium'))
    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 15

    if not pdf_path:
        c = ws.cell(row=4, column=1, value=(
            'Pemeriksaan tidak dijalankan — file PDF sumber tidak tersedia untuk sheet ini.'
        ))
        c.font = Font(italic=True, color='595959')
        return

    # ---- Skor risiko ringkas ----
    n_tinggi = sum(1 for f in findings if f['tingkat'] == 'Tinggi')
    n_sedang = sum(1 for f in findings if f['tingkat'] == 'Sedang')
    n_rendah = sum(1 for f in findings if f['tingkat'] == 'Rendah')

    if n_tinggi > 0:
        skor_label, skor_bg, skor_font = 'RISIKO TINGGI', 'FCE4EC', 'C00000'
    elif n_sedang > 0:
        skor_label, skor_bg, skor_font = 'RISIKO SEDANG', 'FFF2CC', '7D6608'
    else:
        skor_label, skor_bg, skor_font = 'RISIKO RENDAH', 'E2EFDA', '375623'

    score_row = 4
    ws.row_dimensions[score_row].height = 30
    ws.merge_cells(start_row=score_row, start_column=1, end_row=score_row, end_column=9)
    c = ws.cell(row=score_row, column=1,
                value=f'  {skor_label}  —  {len(findings)} temuan  '
                      f'(Tinggi: {n_tinggi}, Sedang: {n_sedang}, Rendah: {n_rendah})')
    c.font      = Font(bold=True, size=13, color=skor_font)
    c.fill      = PatternFill(fill_type='solid', fgColor=skor_bg)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border    = Border(left=Side(style='medium', color=skor_font),
                         right=Side(style='medium', color=skor_font),
                         top=Side(style='medium', color=skor_font),
                         bottom=Side(style='medium', color=skor_font))

    # ---- Ringkasan per kategori ----
    kat_row = score_row + 2
    style_header(ws.cell(row=kat_row, column=1, value='Kategori Indikasi'))
    ws.merge_cells(start_row=kat_row, start_column=1, end_row=kat_row, end_column=4)
    style_header(ws.cell(row=kat_row, column=5, value='Jumlah'))
    ws.merge_cells(start_row=kat_row, start_column=5, end_row=kat_row, end_column=6)
    style_header(ws.cell(row=kat_row, column=7, value='Tingkat Tertinggi'))
    ws.merge_cells(start_row=kat_row, start_column=7, end_row=kat_row, end_column=9)
    ws.row_dimensions[kat_row].height = 22

    per_kategori = {}
    tingkat_order = {'Tinggi': 0, 'Sedang': 1, 'Rendah': 2}
    for f in findings:
        entry = per_kategori.setdefault(f['kategori'], {'jumlah': 0, 'tingkat': f['tingkat']})
        entry['jumlah'] += 1
        if tingkat_order[f['tingkat']] < tingkat_order[entry['tingkat']]:
            entry['tingkat'] = f['tingkat']

    r = kat_row
    for kategori, info in sorted(per_kategori.items(),
                                  key=lambda kv: tingkat_order[kv[1]['tingkat']]):
        r += 1
        bg, font_col = TINGKAT_STYLE[info['tingkat']]
        ws.row_dimensions[r].height = 18
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=kategori)
        style_data(c, align='left')
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        c = ws.cell(row=r, column=5, value=info['jumlah'])
        style_data(c, align='center', bold=True)
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
        c = ws.cell(row=r, column=7, value=info['tingkat'])
        style_data(c, align='center', bold=True, bg_color=bg)
        c.font = Font(bold=True, size=10, color=font_col)

    if not findings:
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        c = ws.cell(row=r, column=1, value='Tidak ditemukan indikasi kejanggalan pada rekening ini.')
        c.font = Font(italic=True, color='375623', bold=True)
        c.alignment = Alignment(horizontal='left', vertical='center')

    # ---- Tabel detail temuan ----
    tbl_row = r + 3
    headers = ['No', 'Kategori', 'Tingkat', 'Bulan', 'Tanggal', 'Halaman',
               'Deskripsi Temuan', 'Detail / Bukti', 'Nilai Terkait (Rp)']
    ws.row_dimensions[tbl_row].height = 26
    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=tbl_row, column=col, value=h))
    ws.freeze_panes = ws.cell(row=tbl_row + 1, column=1)
    ws.auto_filter.ref = f'A{tbl_row}:I{tbl_row}'

    for i, f in enumerate(findings, start=1):
        row = tbl_row + i
        bg, font_col = TINGKAT_STYLE[f['tingkat']]
        ws.row_dimensions[row].height = 30

        vals = [i, f['kategori'], f['tingkat'], f['bulan'], f['tanggal'],
                f['halaman'], f['deskripsi'], f['detail'], f['nilai_rp']]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col, value=v)
            align = 'center' if col in (1, 3, 4, 5, 6) else 'left'
            style_data(c, align=align, bg_color=bg if col == 3 else None)
            if col == 3:
                c.font = Font(bold=True, size=10, color=font_col)
            if col == 9 and v is not None:
                c.number_format = '#,##0'
            if col in (7, 8):
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
